#!/usr/bin/env python3
"""
Excel file checking utilities for the autocheck tool.
"""

import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print("Error: This script requires openpyxl. Install with: pip install openpyxl")
    import sys
    sys.exit(1)


def check_formulas(workbook, example_name):
    """Check for common formula issues in the workbook"""
    issues_found = False
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        for row_idx, row in enumerate(sheet.rows, 1):
            for col_idx, cell in enumerate(row, 1):
                # Check for formula cells
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Check for circular references
                    # This is a basic check - it only catches obvious self-references
                    cell_col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_ref = f"{cell_col_letter}{row_idx}"
                    
                    if cell_ref in cell.value:
                        print(f"[{example_name}] ⚠️ Potential circular reference in {sheet_name}!{cell.coordinate}: {cell.value} references its own cell {cell_ref}")
                        issues_found = True
                    
                    # Check for SUM ranges that might include the formula cell itself
                    sum_match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell.value)
                    if sum_match:
                        start_col, start_row, end_col, end_row = sum_match.groups()
                        start_row, end_row = int(start_row), int(end_row)
                        
                        if (cell_col_letter >= start_col and cell_col_letter <= end_col and 
                            row_idx >= start_row and row_idx <= end_row):
                            print(f"[{example_name}] ⚠️ Formula range includes its own cell in {sheet_name}!{cell.coordinate}: {cell.value}")
                            issues_found = True
                    
                    # Check for null-termination issues (common in the Zig libxlsxwriter wrapper)
                    if cell.value.endswith('\x00'):
                        print(f"[{example_name}] ⚠️ Formula contains null terminator at the end in {sheet_name}!{cell.coordinate}: {cell.value}")
                        issues_found = True
                    
                    # Check for other common formula syntax issues
                    if ':' in cell.value and not re.search(r'[A-Z]+\d+:[A-Z]+\d+', cell.value):
                        print(f"[{example_name}] ⚠️ Potentially malformed range in formula at {sheet_name}!{cell.coordinate}: {cell.value}")
                        issues_found = True
    
    return not issues_found


def check_string_null_termination(workbook, example_name=None):
    """Check for issues with string null termination"""
    issues_found = False
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        for row in sheet.rows:
            for cell in row:
                # Check for string cells with null terminators
                if isinstance(cell.value, str):
                    if '\x00' in cell.value:
                        if example_name:
                            print(f"[{example_name}] ⚠️ Cell {cell.coordinate} contains null character: {repr(cell.value)}")
                        else:
                            print(f"⚠️ Cell {cell.coordinate} contains null character: {repr(cell.value)}")
                        issues_found = True
                    
                    # Check for truncated strings (potential null termination issues)
                    if cell.value.endswith('...') or cell.value.endswith('…'):
                        if example_name:
                            print(f"[{example_name}] ⚠️ Cell {cell.coordinate} might be truncated: {cell.value}")
                        else:
                            print(f"⚠️ Cell {cell.coordinate} might be truncated: {cell.value}")
                        issues_found = True
    
    return not issues_found


def check_xml_content(example_name):
    """Check the XML content of the xlsx file for encoding issues in memory"""
    generated_file = Path(f"{example_name}.xlsx")
    
    try:
        with zipfile.ZipFile(generated_file, 'r') as xlsx_zip:
            # Get list of all XML files in the workbook
            sheet_files = [name for name in xlsx_zip.namelist() 
                         if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')]
            
            for sheet_file in sheet_files:
                with xlsx_zip.open(sheet_file) as file:
                    try:
                        tree = ET.parse(file)
                        root = tree.getroot()
                        
                        # Look for formula elements
                        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        formulas = root.findall(".//s:f", ns)
                        
                        for formula in formulas:
                            formula_text = formula.text or ""
                            
                            # Check if formula has proper XML escaping
                            if '<' in formula_text or '>' in formula_text or '&' in formula_text:
                                print(f"[{example_name}] ⚠️ Formula contains XML special characters that may need escaping: {formula_text}")
                            
                            # Check if formula is truncated or malformed
                            if formula_text.startswith('=') and len(formula_text) < 3:
                                print(f"[{example_name}] ⚠️ Formula seems truncated or malformed: {formula_text}")
                            
                            # Check for null characters in XML (which could indicate issues with Zig's string handling)
                            if '\x00' in formula_text:
                                print(f"[{example_name}] ⚠️ Formula contains null characters which may cause issues: {formula_text}")
                        
                        # Check for other string content (similarly might have null termination issues)
                        cells = root.findall(".//s:c/s:v", ns)
                        for cell in cells:
                            if cell.text and '\x00' in cell.text:
                                print(f"[{example_name}] ⚠️ Cell value contains null characters: {repr(cell.text)}")
                        
                    except ET.ParseError as e:
                        print(f"[{example_name}] ❌ XML parsing error in {sheet_file}: {e}")
                        return False
            
            # Check for specific string table entries (shared strings)
            if 'xl/sharedStrings.xml' in xlsx_zip.namelist():
                with xlsx_zip.open('xl/sharedStrings.xml') as file:
                    try:
                        tree = ET.parse(file)
                        root = tree.getroot()
                        
                        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        strings = root.findall(".//s:t", ns)
                        
                        for string in strings:
                            string_text = string.text or ""
                            
                            # Check for null characters in shared strings
                            if '\x00' in string_text:
                                print(f"[{example_name}] ⚠️ Shared string contains null characters: {repr(string_text)}")
                            
                            # Check for potentially malformed strings
                            if string_text.endswith('...') or string_text.endswith('…'):
                                print(f"[{example_name}] ⚠️ Shared string might be truncated: {string_text}")
                    
                    except ET.ParseError as e:
                        print(f"[{example_name}] ❌ XML parsing error in sharedStrings.xml: {e}")
                        return False
    
    except zipfile.BadZipFile:
        print(f"[{example_name}] ❌ File is not a valid ZIP/XLSX file: {generated_file.name}")
        return False
    
    return True


def check_binary_compatibility(example_name, reference_dir):
    """Check for binary compatibility issues that might not be visible in the content"""
    generated_file = Path(f"{example_name}.xlsx")
    reference_file = reference_dir / f"{example_name}.xlsx"
    
    if not reference_file.exists():
        print(f"[{example_name}] ⚠️ Reference file not found: {reference_file}")
        return False
    
    try:
        has_differences = False
        
        # Compare file sizes - significant differences might indicate issues
        gen_size = generated_file.stat().st_size
        ref_size = reference_file.stat().st_size
        size_diff_percent = abs(gen_size - ref_size) / max(gen_size, ref_size) * 100
        
        if size_diff_percent > 10:  # More than 10% size difference
            print(f"[{example_name}] ⚠️ File size differs significantly: {gen_size} vs {ref_size} bytes ({size_diff_percent:.2f}% difference)")
            has_differences = True
        
        # Check internal file structure using zipfile
        with zipfile.ZipFile(generated_file, 'r') as gen_zip, zipfile.ZipFile(reference_file, 'r') as ref_zip:
            gen_files = set(gen_zip.namelist())
            ref_files = set(ref_zip.namelist())
            
            # Check for missing files
            missing_files = ref_files - gen_files
            if missing_files:
                print(f"[{example_name}] ⚠️ Generated file is missing these internal files: {missing_files}")
                has_differences = True
            
            # Check for extra files
            extra_files = gen_files - ref_files
            if extra_files:
                print(f"[{example_name}] ⚠️ Generated file has these extra internal files: {extra_files}")
                has_differences = True
            
            # Compare contents of important files
            for file_name in ['xl/workbook.xml', 'xl/styles.xml']:
                if file_name in gen_files and file_name in ref_files:
                    with gen_zip.open(file_name) as gen_file, ref_zip.open(file_name) as ref_file:
                        gen_content = gen_file.read()
                        ref_content = ref_file.read()
                        
                        if gen_content != ref_content:
                            print(f"[{example_name}] ⚠️ Content of {file_name} differs")
                            has_differences = True
        
        return not has_differences
    
    except Exception as e:
        print(f"[{example_name}] ❌ Error checking binary compatibility: {e}")
        return False


def check_row_visibility(example_name, reference_dir):
    """Check that row visibility states match between generated and reference files"""
    generated_file = Path(f"{example_name}.xlsx")
    reference_file = reference_dir / f"{example_name}.xlsx"
    
    if not reference_file.exists():
        print(f"[{example_name}] ⚠️ Reference file not found: {reference_file}")
        return False
    
    try:
        has_differences = False
        
        # Load both workbooks
        gen_wb = openpyxl.load_workbook(generated_file)
        ref_wb = openpyxl.load_workbook(reference_file)
        
        # Compare each sheet
        for sheet_name in ref_wb.sheetnames:
            if sheet_name not in gen_wb.sheetnames:
                print(f"[{example_name}] ⚠️ Generated file is missing sheet: {sheet_name}")
                has_differences = True
                continue
            
            ref_sheet = ref_wb[sheet_name]
            gen_sheet = gen_wb[sheet_name]
            
            # Get the maximum row number to check
            max_row = max(ref_sheet.max_row, gen_sheet.max_row)
            
            # Check each row's visibility
            for row in range(1, max_row + 1):
                # Get row dimensions, defaulting to visible if not set
                ref_dim = ref_sheet.row_dimensions.get(row)
                gen_dim = gen_sheet.row_dimensions.get(row)
                
                ref_hidden = ref_dim.hidden if ref_dim and ref_dim.hidden is not None else False
                gen_hidden = gen_dim.hidden if gen_dim and gen_dim.hidden is not None else False
                
                if ref_hidden != gen_hidden:
                    print(f"[{example_name}] ⚠️ Row visibility mismatch in sheet '{sheet_name}' at row {row}:")
                    print(f"  Reference: {'hidden' if ref_hidden else 'visible'}")
                    print(f"  Generated: {'hidden' if gen_hidden else 'visible'}")
                    has_differences = True
                
                # Also check row heights if they differ significantly
                ref_height = ref_dim.height if ref_dim and ref_dim.height is not None else 15
                gen_height = gen_dim.height if gen_dim and gen_dim.height is not None else 15
                
                if abs(ref_height - gen_height) > 0.1:  # Allow small floating point differences
                    print(f"[{example_name}] ⚠️ Row height mismatch in sheet '{sheet_name}' at row {row}:")
                    print(f"  Reference: {ref_height}")
                    print(f"  Generated: {gen_height}")
                    has_differences = True
        
        # Check for extra sheets in generated file
        for sheet_name in gen_wb.sheetnames:
            if sheet_name not in ref_wb.sheetnames:
                print(f"[{example_name}] ⚠️ Generated file has extra sheet: {sheet_name}")
                has_differences = True
        
        return not has_differences
    
    except Exception as e:
        print(f"[{example_name}] ❌ Error checking row visibility: {e}")
        return False 