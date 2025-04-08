#!/usr/bin/env python3
"""
File comparison utilities for the autocheck tool.
"""

from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

passed_autocheck_file = "autochecked"

def get_relative_path(path, project_root):
    """Convert a path to be relative to the project root if possible"""
    try:
        return path.relative_to(project_root)
    except ValueError:
        # If the path can't be made relative to PROJECT_ROOT, return as is
        return path


def compare_with_reference(example_name, reference_dir, results_dir, project_root, quiet=False, ignore_styles=False):
    """Compare the generated Excel file with the reference file"""
    generated_file = Path(f"{example_name}.xlsx")
    generated_macro_file = Path(f"{example_name}.xlsm")
    reference_file = reference_dir / f"{example_name}.xlsx"
    reference_macro_file = reference_dir / f"{example_name}.xlsm"
    
    # Use the macro files if they exist, otherwise use the regular files
    file_to_check = generated_macro_file if generated_macro_file.exists() else generated_file
    ref_to_check = reference_macro_file if reference_macro_file.exists() else reference_file
    
    if not ref_to_check.exists():
        print(f"[{example_name}] ⚠️ Reference file not found: {ref_to_check}")
        return False
    
    try:
        has_differences = False
        
        # Load both workbooks
        gen_wb = openpyxl.load_workbook(file_to_check)
        ref_wb = openpyxl.load_workbook(ref_to_check)
        
        # Compare each sheet
        for sheet_name in ref_wb.sheetnames:
            if sheet_name not in gen_wb.sheetnames:
                print(f"[{example_name}] ⚠️ Generated file is missing sheet: {sheet_name}")
                has_differences = True
                continue
            
            ref_sheet = ref_wb[sheet_name]
            gen_sheet = gen_wb[sheet_name]
            
            # Skip chartsheets as they don't have rows/cells
            if isinstance(ref_sheet, openpyxl.chartsheet.Chartsheet) or isinstance(gen_sheet, openpyxl.chartsheet.Chartsheet):
                continue
            
            # Get the maximum row and column numbers to check
            max_row = max(ref_sheet.max_row, gen_sheet.max_row)
            max_col = max(ref_sheet.max_column, gen_sheet.max_column)
            
            # Compare cell values
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ref_cell = ref_sheet.cell(row=row, column=col)
                    gen_cell = gen_sheet.cell(row=row, column=col)
                    
                    # Compare values
                    if ref_cell.value != gen_cell.value:
                        print(f"[{example_name}] ⚠️ Value mismatch in sheet '{sheet_name}' at {ref_cell.coordinate}:")
                        print(f"  Reference: {ref_cell.value}")
                        print(f"  Generated: {gen_cell.value}")
                        has_differences = True
                    
                    # Only compare styles if explicitly requested and not a chartsheet example
                    if not ignore_styles and example_name != "chartsheet":
                        if ref_cell.font != gen_cell.font:
                            print(f"[{example_name}] ⚠️ Font mismatch in sheet '{sheet_name}' at {ref_cell.coordinate}")
                            has_differences = True
                        if ref_cell.fill != gen_cell.fill:
                            print(f"[{example_name}] ⚠️ Fill mismatch in sheet '{sheet_name}' at {ref_cell.coordinate}")
                            has_differences = True
                        if ref_cell.border != gen_cell.border:
                            print(f"[{example_name}] ⚠️ Border mismatch in sheet '{sheet_name}' at {ref_cell.coordinate}")
                            has_differences = True
                        if ref_cell.alignment != gen_cell.alignment:
                            print(f"[{example_name}] ⚠️ Alignment mismatch in sheet '{sheet_name}' at {ref_cell.coordinate}")
                            has_differences = True
        
        # Check for extra sheets in generated file
        for sheet_name in gen_wb.sheetnames:
            if sheet_name not in ref_wb.sheetnames:
                print(f"[{example_name}] ⚠️ Generated file has extra sheet: {sheet_name}")
                has_differences = True
        
        # Report any differences found
        if has_differences:
            if not quiet:
                print("\n❌ Generated file has differences from reference file")
        else:
            if not quiet:
                print("✅ Generated file matches reference file content" + 
                     (" (ignoring styles)" if ignore_styles or example_name == "chartsheet" else ""))
        
        # Create results directory if it doesn't exist
        example_results_dir = results_dir / example_name
        example_results_dir.mkdir(parents=True, exist_ok=True)
        
        # Create or remove autochecked file based on results
        autochecked_file = example_results_dir / passed_autocheck_file
        if not has_differences:
            autochecked_file.touch()
            if not quiet:
                print(f"✅ Created autochecked file at {get_relative_path(autochecked_file, project_root)}")
        else:
            if autochecked_file.exists():
                autochecked_file.unlink()
            if not quiet:
                print(f"❌ Removed autochecked file due to differences")
        
        return not has_differences
    
    except Exception as e:
        print(f"[{example_name}] ❌ Error comparing with reference: {e}")
        return False 