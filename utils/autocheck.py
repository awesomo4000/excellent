#!/usr/bin/env python3
"""
Automated checks for Excel files to identify potential issues before manual verification.
This tool examines Excel files for common problems like formula warnings, missing data,
or encoding issues.

Common usage:
  python3 utils/autocheck.py example_name --build --run        # Build, run and check
  python3 utils/autocheck.py example_name --ignore-styles      # Ignore style differences
  python3 utils/autocheck.py --all                             # Check all examples
  python3 utils/autocheck.py --list-broken                     # List known broken examples
"""

import os
import sys
import argparse
from pathlib import Path
import openpyxl
import traceback

# Add the project root to sys.path to allow imports to work both when run as a module 
# and directly
current_dir = Path(__file__).parent
project_root = current_dir.parent
sys.path.insert(0, str(project_root))

# Import local modules
try:
    # When run as module (python -m utils.autocheck)
    from utils.excel_checks import (
        check_formulas,
        check_string_null_termination,
        check_xml_content,
        check_binary_compatibility,
        check_row_visibility
    )
    from utils.file_comparison import compare_with_reference, get_relative_path
    from utils.example_runner import build_example, run_example
except ModuleNotFoundError:
    # When run directly (python utils/autocheck.py)
    from excel_checks import (
        check_formulas,
        check_string_null_termination,
        check_xml_content,
        check_binary_compatibility,
        check_row_visibility
    )
    from file_comparison import compare_with_reference, get_relative_path
    from example_runner import build_example, run_example

# Set up paths
PROJECT_ROOT = Path(__file__).parent.parent
EXAMPLES_DIR = PROJECT_ROOT / "examples"
REFERENCE_DIR = PROJECT_ROOT / "testing" / "reference-xls"
RESULTS_DIR = PROJECT_ROOT / "testing" / "results"
BROKEN_FILE = PROJECT_ROOT / "testing" / ".broken"


def load_broken_examples():
    """Load list of known broken examples from testing/.broken file"""
    broken_examples = set()
    
    if BROKEN_FILE.exists():
        with open(BROKEN_FILE, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    broken_examples.add(line)
    
    return broken_examples


def is_broken_example(example_name, broken_examples=None):
    """Check if example is in the list of known broken examples"""
    if broken_examples is None:
        broken_examples = load_broken_examples()
    
    return example_name in broken_examples


def check_single_example(example_name, args):
    """Run checks on a single example"""
    # Check if example is known to be broken
    broken_examples = load_broken_examples()
    is_broken = is_broken_example(example_name, broken_examples)
    
    if is_broken and not args.force:
        print(f"{example_name} ⚠️ is listed in testing/.broken as a known broken example")
        print("Use --force to check it anyway")
        return True  # Return success for broken examples unless forced
    
    # Check if example exists, unless in file-only mode
    if not args.file_only:
        example_file = EXAMPLES_DIR / f"{example_name}.zig"
        if not example_file.exists():
            print(f"❌ Example file not found: {get_relative_path(example_file, PROJECT_ROOT)}")
            return False
    
    # Build if requested
    if args.build:
        if not build_example(example_name, PROJECT_ROOT):
            return is_broken  # Return success for broken examples
    
    # Run if requested
    if args.run:
        if not run_example(example_name, PROJECT_ROOT):
            return is_broken  # Return success for broken examples
    
    # Look for the Excel file
    excel_file = Path(f"{example_name}.xlsx")
    excel_macro_file = Path(f"{example_name}.xlsm")
    if not excel_file.exists() and not excel_macro_file.exists():
        print(f"❌ Excel file not found: {excel_file} or {excel_macro_file}")
        print("Run with --run option to generate it")
        return is_broken  # Return success for broken examples
    
    # Use the macro file if it exists, otherwise use the regular file
    excel_file = excel_macro_file if excel_macro_file.exists() else excel_file
    
    print(f"\n=== Checking Excel file: {excel_file} ===\n")
    
    try:
        # Load workbook for checks
        workbook = openpyxl.load_workbook(excel_file)
        
        # Run checks
        print(f"[{example_name}] Checking formulas...")
        formula_check = check_formulas(workbook, example_name)
        
        print(f"[{example_name}] Checking string null-termination...")
        string_check = check_string_null_termination(workbook, example_name)
        
        print(f"[{example_name}] Checking XML content...")
        xml_check = check_xml_content(example_name)
        
        print(f"[{example_name}] Checking binary compatibility...")
        binary_check = check_binary_compatibility(example_name, REFERENCE_DIR)
        
        print(f"[{example_name}] Checking row visibility...")
        visibility_check = check_row_visibility(example_name, REFERENCE_DIR)
        
        print(f"[{example_name}] Comparing with reference file...")
        content_check = compare_with_reference(
            example_name, 
            REFERENCE_DIR, 
            RESULTS_DIR, 
            PROJECT_ROOT, 
            ignore_styles=args.ignore_styles
        )
        
        # Summary
        print(f"\n=== Check Summary for {example_name} ===")
        print(f"Formula Check: {'✅ PASSED' if formula_check else '❌ FAILED'}")
        print(f"String Null-Termination: {'✅ PASSED' if string_check else '❌ FAILED'}")
        print(f"XML Check: {'✅ PASSED' if xml_check else '❌ FAILED'}")
        print(f"Binary Compatibility: {'✅ PASSED' if binary_check else '❌ FAILED'}")
        print(f"Row Visibility: {'✅ PASSED' if visibility_check else '❌ FAILED'}")
        print(f"Content Check: {'✅ PASSED' if content_check else '❌ FAILED'}")
        
        all_passed = formula_check and string_check and xml_check and binary_check and visibility_check and content_check
        
        if is_broken and not all_passed:
            print(f"\n{example_name} ⚠️ Known broken example failed checks as expected.")
            return True  # Return success for broken examples that fail
        elif is_broken and all_passed:
            print(f"\n{example_name} ⚠️ This example is listed as broken but all checks passed!")
            print("Consider removing it from testing/.broken")
            return False  # Fail when a "broken" example passes all checks
        elif all_passed:
            print(f"\n{example_name} ✅ All checks passed! The file should pass manual verification.")
            return True
        else:
            print(f"\n{example_name} ⚠️ Some checks failed. Review issues before manual verification.")
            return False
            
    except Exception as e:
        print(f"{example_name} ❌ Error checking Excel file: {e}")
        traceback.print_exc()
        return is_broken  # Return success for broken examples


def check_all_examples(args):
    """Run checks on all examples"""
    # Get all example files
    example_files = list(EXAMPLES_DIR.glob("*.zig"))
    failed_examples = []
    broken_examples = load_broken_examples()
    
    for example_file in example_files:
        example_name = example_file.stem
        if example_name == "status":  # Skip status binary
            continue
            
        # Skip broken examples unless forced
        if is_broken_example(example_name, broken_examples) and not args.force:
            print(f"{example_name} [BROKEN] ⚠️ Skipping as known broken example")
            continue
            
        # Build if requested
        if args.build:
            print(f"Building {example_name} ... ", end="", flush=True)
            if not build_example(example_name, PROJECT_ROOT, quiet=True):
                if is_broken_example(example_name, broken_examples):
                    print("✅ [BROKEN] Failed as expected")
                    continue
                print("❌")
                failed_examples.append((example_name, "Failed to build example"))
                continue
            # Print success for build-only mode
            if not args.run:
                print("✅")
                continue
            
        # Run if requested
        if args.run:
            if not run_example(example_name, PROJECT_ROOT, quiet=True):
                if is_broken_example(example_name, broken_examples):
                    print(f"{example_name} [BROKEN] ✅ Failed to run as expected")
                    continue
                failed_examples.append((example_name, "Failed to run example"))
                continue
            
        # Only check Excel files if we're not in build-only mode
        if not (args.build and not args.run):
            # Check the Excel file
            excel_file = Path(f"{example_name}.xlsx")
            excel_macro_file = Path(f"{example_name}.xlsm")
            if not excel_file.exists() and not excel_macro_file.exists():
                if is_broken_example(example_name, broken_examples):
                    print(f"{example_name} [BROKEN] ✅ Excel file not generated as expected")
                    continue
                failed_examples.append((example_name, "Excel file not generated"))
                continue
            
            try:
                # Use the macro file if it exists, otherwise use the regular file
                file_to_check = excel_macro_file if excel_macro_file.exists() else excel_file
                # Load workbook for checks
                workbook = openpyxl.load_workbook(file_to_check)
                
                # Run checks
                formula_check = check_formulas(workbook, example_name)
                string_check = check_string_null_termination(workbook, example_name)
                xml_check = check_xml_content(example_name)
                binary_check = check_binary_compatibility(example_name, REFERENCE_DIR)
                visibility_check = check_row_visibility(example_name, REFERENCE_DIR)
                content_check = compare_with_reference(
                    example_name, 
                    REFERENCE_DIR, 
                    RESULTS_DIR, 
                    PROJECT_ROOT, 
                    quiet=True, 
                    ignore_styles=args.ignore_styles
                )
                
                all_passed = formula_check and string_check and xml_check and binary_check and visibility_check and content_check
                
                if is_broken_example(example_name, broken_examples):
                    if all_passed:
                        print(f"{example_name} [BROKEN] ⚠️ Unexpectedly passed all checks")
                        failed_examples.append((example_name, "Broken example passed all checks"))
                    else:
                        print(f"{example_name} [BROKEN] ✅ Failed checks as expected")
                elif all_passed:
                    print(f"{example_name} ✅")
                else:
                    failed_examples.append((example_name, "One or more checks failed"))
                    # Re-run checks with verbose output to show details
                    print(f"\nDetailed output for {example_name}:")
                    try:
                        workbook = openpyxl.load_workbook(Path(f"{example_name}.xlsx"))
                        print(f"\n[{example_name}] Checking formulas...")
                        check_formulas(workbook, example_name)
                        print(f"\n[{example_name}] Checking string null-termination...")
                        check_string_null_termination(workbook, example_name)
                        print(f"\n[{example_name}] Checking XML content...")
                        check_xml_content(example_name)
                        print(f"\n[{example_name}] Checking binary compatibility...")
                        check_binary_compatibility(example_name, REFERENCE_DIR)
                        print(f"\n[{example_name}] Checking row visibility...")
                        check_row_visibility(example_name, REFERENCE_DIR)
                        print(f"\n[{example_name}] Comparing with reference file...")
                        compare_with_reference(
                            example_name, 
                            REFERENCE_DIR, 
                            RESULTS_DIR, 
                            PROJECT_ROOT, 
                            ignore_styles=args.ignore_styles
                        )
                    except Exception as e:
                        print(f"Error during detailed check: {e}")
                        traceback.print_exc()
                
            except Exception as e:
                if is_broken_example(example_name, broken_examples):
                    print(f"{example_name} [BROKEN] ✅ Error occurred as expected: {e}")
                    continue
                failed_examples.append((example_name, f"Error checking Excel file: {e}"))
    
    # Print detailed failure information if any
    if failed_examples:
        print("\n=== Failed Examples ===")
        for example_name, error in failed_examples:
            print(f"\n{example_name} ❌ {error}")
        
        return False
    else:
        print("\n✅ All examples passed!")
        return True


def list_broken_examples():
    """List examples marked as broken in testing/.broken"""
    broken_examples = load_broken_examples()
    
    if not broken_examples:
        print("No examples are currently marked as broken.")
        return
    
    print("=== Known Broken Examples ===")
    for example in sorted(broken_examples):
        print(f"- {example}")


def main():
    parser = argparse.ArgumentParser(description="Check Excel files for common issues before verification")
    parser.add_argument("example", nargs="?", help="Name of the example to check (without .zig extension)")
    parser.add_argument("--build", action="store_true", help="Build the example before checking")
    parser.add_argument("--run", action="store_true", help="Run the example to generate the Excel file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed information about checks")
    parser.add_argument("--file-only", action="store_true", help="Skip example file check, just check the Excel file")
    parser.add_argument("--all", action="store_true", help="Check all examples (except status)")
    parser.add_argument("--ignore-styles", action="store_true", help="Ignore style differences in comparison")
    parser.add_argument("--force", "-f", action="store_true", help="Force checking of known broken examples")
    parser.add_argument("--list-broken", action="store_true", help="List examples marked as broken")
    
    args = parser.parse_args()
    
    if args.list_broken:
        list_broken_examples()
        return 0
    
    if args.all:
        if args.example:
            print("Error: Cannot specify both --all and an example name")
            return 1
        
        return 0 if check_all_examples(args) else 1
    
    if not args.example:
        print("Error: Must specify either an example name or --all or --list-broken")
        return 1
    
    return 0 if check_single_example(args.example, args) else 1


if __name__ == "__main__":
    sys.exit(main()) 