#!/usr/bin/env python3
"""
Automated checks for Excel files to identify potential issues before manual verification.
This tool examines Excel files for common problems like formula warnings, missing data,
or encoding issues.
"""

import os
import sys
import argparse
from pathlib import Path
import subprocess
import tempfile
import re
import xml.etree.ElementTree as ET
import zipfile
import io

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print("Error: This script requires openpyxl. Install with: pip install openpyxl")
    sys.exit(1)

# Set up paths
PROJECT_ROOT = Path(__file__).parent.parent
EXAMPLES_DIR = PROJECT_ROOT / "examples"
REFERENCE_DIR = PROJECT_ROOT / "testing" / "reference-xls"
OUTPUT_DIR = PROJECT_ROOT / "testing" / "test-output-xls"


def get_relative_path(path):
    """Convert a path to be relative to the project root if possible"""
    try:
        return path.relative_to(PROJECT_ROOT)
    except ValueError:
        # If the path can't be made relative to PROJECT_ROOT, return as is
        return path


def build_example(example_name):
    """Build the example and return True if successful"""
    print(f"Building example: {example_name}")
    cmd = ["zig", "build", "examples", f"-Dexample={example_name}"]
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode != 0:
        print(f"❌ Build failed for {example_name}")
        print(result.stderr)
        return False
    
    print(f"✅ Build successful for {example_name}")
    return True


def run_example(example_name):
    """Run the example to generate the Excel file"""
    print(f"Running example to generate Excel file: {example_name}.xlsx")
    example_bin = PROJECT_ROOT / "zig-out" / "bin" / example_name
    
    if not example_bin.exists():
        print(f"❌ Executable not found at {get_relative_path(example_bin)}")
        return False
    
    result = subprocess.run([str(example_bin)], capture_output=True, text=True)
    
    if result.returncode != 0:
        print(f"❌ Example execution failed for {example_name}")
        print(result.stderr)
        return False
    
    generated_file = Path(f"{example_name}.xlsx")
    if not generated_file.exists():
        print(f"❌ Excel file not generated: {generated_file}")
        return False
    
    print(f"✅ Excel file generated: {generated_file}")
    return True


def check_formulas(workbook, example_name):
    """Check for common formula issues in the workbook"""
    issues_found = False
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        for row_idx, row in enumerate(sheet.rows, 1):
            for col_idx, cell in enumerate(row, 1):
                # Check for formula cells
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"Found formula at {sheet_name}!{cell.coordinate}: {cell.value}")
                    
                    # Check for circular references
                    # This is a basic check - it only catches obvious self-references
                    cell_col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_ref = f"{cell_col_letter}{row_idx}"
                    
                    if cell_ref in cell.value:
                        print(f"⚠️ Potential circular reference: {cell.value} references its own cell {cell_ref}")
                        issues_found = True
                    
                    # Check for SUM ranges that might include the formula cell itself
                    sum_match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell.value)
                    if sum_match:
                        start_col, start_row, end_col, end_row = sum_match.groups()
                        start_row, end_row = int(start_row), int(end_row)
                        
                        if (cell_col_letter >= start_col and cell_col_letter <= end_col and 
                            row_idx >= start_row and row_idx <= end_row):
                            print(f"⚠️ Formula range includes its own cell: {cell.value}")
                            issues_found = True
                    
                    # Check for null-termination issues (common in the Zig libxlsxwriter wrapper)
                    if cell.value.endswith('\x00'):
                        print(f"⚠️ Formula contains null terminator at the end: {cell.value}")
                        issues_found = True
                    
                    # Check for other common formula syntax issues
                    if ':' in cell.value and not re.search(r'[A-Z]+\d+:[A-Z]+\d+', cell.value):
                        print(f"⚠️ Potentially malformed range in formula: {cell.value}")
                        issues_found = True
    
    return not issues_found


def check_string_null_termination(workbook):
    """Check for issues with string null termination"""
    issues_found = False
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        for row in sheet.rows:
            for cell in row:
                # Check for string cells with null terminators
                if isinstance(cell.value, str):
                    if '\x00' in cell.value:
                        print(f"⚠️ Cell {cell.coordinate} contains null character: {repr(cell.value)}")
                        issues_found = True
                    
                    # Check for truncated strings (potential null termination issues)
                    if cell.value.endswith('...') or cell.value.endswith('…'):
                        print(f"⚠️ Cell {cell.coordinate} might be truncated: {cell.value}")
                        issues_found = True
    
    return not issues_found


def compare_with_reference(example_name):
    """Compare generated Excel file with reference file"""
    generated_file = Path(f"{example_name}.xlsx")
    reference_file = REFERENCE_DIR / f"{example_name}.xlsx"
    
    if not reference_file.exists():
        print(f"⚠️ Reference file not found: {get_relative_path(reference_file)}")
        return False
    
    try:
        gen_wb = openpyxl.load_workbook(generated_file)
        ref_wb = openpyxl.load_workbook(reference_file)
        
        # Compare sheet names
        if gen_wb.sheetnames != ref_wb.sheetnames:
            print(f"⚠️ Sheet names differ: {gen_wb.sheetnames} vs {ref_wb.sheetnames}")
            return False
        
        # Compare cell contents in each sheet
        for sheet_name in gen_wb.sheetnames:
            gen_sheet = gen_wb[sheet_name]
            ref_sheet = ref_wb[sheet_name]
            
            # Find the max dimensions to iterate through
            max_row = max(gen_sheet.max_row, ref_sheet.max_row)
            max_col = max(gen_sheet.max_column, ref_sheet.max_column)
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    gen_cell = gen_sheet.cell(row=row, column=col)
                    ref_cell = ref_sheet.cell(row=row, column=col)
                    
                    if gen_cell.value != ref_cell.value:
                        print(f"⚠️ Cell value mismatch at {sheet_name}!{gen_cell.coordinate}: "
                              f"'{gen_cell.value}' vs '{ref_cell.value}'")
                        return False
        
        print("✅ Generated file matches reference file content")
        return True
        
    except InvalidFileException as e:
        print(f"❌ Error opening Excel file: {e}")
        return False


def check_xml_content(example_name):
    """Check the XML content of the xlsx file for encoding issues in memory"""
    generated_file = Path(f"{example_name}.xlsx")
    
    try:
        with zipfile.ZipFile(generated_file, 'r') as xlsx_zip:
            # Get list of all XML files in the workbook
            sheet_files = [name for name in xlsx_zip.namelist() 
                          if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')]
            
            for sheet_file in sheet_files:
                with xlsx_zip.open(sheet_file) as file:
                    try:
                        tree = ET.parse(file)
                        root = tree.getroot()
                        
                        # Look for formula elements
                        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        formulas = root.findall(".//s:f", ns)
                        
                        for formula in formulas:
                            formula_text = formula.text or ""
                            print(f"Found formula in XML: {formula_text}")
                            
                            # Check if formula has proper XML escaping
                            if '<' in formula_text or '>' in formula_text or '&' in formula_text:
                                print("⚠️ Formula contains XML special characters that may need escaping")
                            
                            # Check if formula is truncated or malformed
                            if formula_text.startswith('=') and len(formula_text) < 3:
                                print("⚠️ Formula seems truncated or malformed")
                            
                            # Check for null characters in XML (which could indicate issues with Zig's string handling)
                            if '\x00' in formula_text:
                                print("⚠️ Formula contains null characters which may cause issues")
                        
                        # Check for other string content (similarly might have null termination issues)
                        cells = root.findall(".//s:c/s:v", ns)
                        for cell in cells:
                            if cell.text and '\x00' in cell.text:
                                print(f"⚠️ Cell value contains null characters: {repr(cell.text)}")
                        
                    except ET.ParseError as e:
                        print(f"❌ XML parsing error in {sheet_file}: {e}")
                        return False
            
            # Check for specific string table entries (shared strings)
            if 'xl/sharedStrings.xml' in xlsx_zip.namelist():
                with xlsx_zip.open('xl/sharedStrings.xml') as file:
                    try:
                        tree = ET.parse(file)
                        root = tree.getroot()
                        
                        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        strings = root.findall(".//s:t", ns)
                        
                        for string in strings:
                            string_text = string.text or ""
                            
                            # Check for null characters in shared strings
                            if '\x00' in string_text:
                                print(f"⚠️ Shared string contains null characters: {repr(string_text)}")
                            
                            # Check for potentially malformed strings
                            if string_text.endswith('...') or string_text.endswith('…'):
                                print(f"⚠️ Shared string might be truncated: {string_text}")
                    
                    except ET.ParseError as e:
                        print(f"❌ XML parsing error in sharedStrings.xml: {e}")
                        return False
    
    except zipfile.BadZipFile:
        print(f"❌ File is not a valid ZIP/XLSX file: {generated_file}")
        return False
    
    return True


def check_binary_compatibility(example_name):
    """Check for binary compatibility issues that might not be visible in the content"""
    generated_file = Path(f"{example_name}.xlsx")
    reference_file = REFERENCE_DIR / f"{example_name}.xlsx"
    
    if not reference_file.exists():
        print(f"⚠️ Reference file not found: {get_relative_path(reference_file)}")
        return False
    
    try:
        # Compare file sizes - significant differences might indicate issues
        gen_size = generated_file.stat().st_size
        ref_size = reference_file.stat().st_size
        size_diff_percent = abs(gen_size - ref_size) / max(gen_size, ref_size) * 100
        
        if size_diff_percent > 10:  # More than 10% size difference
            print(f"⚠️ File size differs significantly: {gen_size} vs {ref_size} bytes ({size_diff_percent:.2f}% difference)")
        
        # Check internal file structure using zipfile
        with zipfile.ZipFile(generated_file, 'r') as gen_zip, zipfile.ZipFile(reference_file, 'r') as ref_zip:
            gen_files = set(gen_zip.namelist())
            ref_files = set(ref_zip.namelist())
            
            # Check for missing files
            missing_files = ref_files - gen_files
            if missing_files:
                print(f"⚠️ Generated file is missing these internal files: {missing_files}")
            
            # Check for extra files
            extra_files = gen_files - ref_files
            if extra_files:
                print(f"⚠️ Generated file has these extra internal files: {extra_files}")
            
            # Compare contents of important files
            for file_name in ['xl/workbook.xml', 'xl/styles.xml']:
                if file_name in gen_files and file_name in ref_files:
                    with gen_zip.open(file_name) as gen_file, ref_zip.open(file_name) as ref_file:
                        gen_content = gen_file.read()
                        ref_content = ref_file.read()
                        
                        if gen_content != ref_content:
                            print(f"⚠️ Content of {file_name} differs")
    
    except Exception as e:
        print(f"❌ Error checking binary compatibility: {e}")
        return False
    
    return True


def main():
    parser = argparse.ArgumentParser(description="Check Excel files for common issues before verification")
    parser.add_argument("example", help="Name of the example to check (without .zig extension)")
    parser.add_argument("--build", action="store_true", help="Build the example before checking")
    parser.add_argument("--run", action="store_true", help="Run the example to generate the Excel file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed information about checks")
    parser.add_argument("--file-only", action="store_true", help="Skip example file check, just check the Excel file")
    
    args = parser.parse_args()
    example_name = args.example
    
    # Check if example exists, unless in file-only mode
    if not args.file_only:
        example_file = EXAMPLES_DIR / f"{example_name}.zig"
        if not example_file.exists():
            print(f"❌ Example file not found: {get_relative_path(example_file)}")
            return 1
    
    # Build if requested
    if args.build:
        if not build_example(example_name):
            return 1
    
    # Run if requested
    if args.run:
        if not run_example(example_name):
            return 1
    
    # Look for the Excel file
    excel_file = Path(f"{example_name}.xlsx")
    if not excel_file.exists():
        print(f"❌ Excel file not found: {excel_file}")
        print("Run with --run option to generate it")
        return 1
    
    print(f"\n=== Checking Excel file: {excel_file} ===\n")
    
    try:
        # Load workbook for checks
        workbook = openpyxl.load_workbook(excel_file)
        
        # Run checks
        print("Checking formulas...")
        formula_check = check_formulas(workbook, example_name)
        
        print("Checking string null-termination...")
        string_check = check_string_null_termination(workbook)
        
        print("Checking XML content...")
        xml_check = check_xml_content(example_name)
        
        print("Checking binary compatibility...")
        binary_check = check_binary_compatibility(example_name)
        
        print("Comparing with reference file...")
        content_check = compare_with_reference(example_name)
        
        # Summary
        print("\n=== Check Summary ===")
        print(f"Formula Check: {'✅ PASSED' if formula_check else '❌ FAILED'}")
        print(f"String Null-Termination: {'✅ PASSED' if string_check else '❌ FAILED'}")
        print(f"XML Check: {'✅ PASSED' if xml_check else '❌ FAILED'}")
        print(f"Binary Compatibility: {'✅ PASSED' if binary_check else '❌ FAILED'}")
        print(f"Content Check: {'✅ PASSED' if content_check else '❌ FAILED'}")
        
        if formula_check and string_check and xml_check and binary_check and content_check:
            print("\n✅ All checks passed! The file should pass manual verification.")
            return 0
        else:
            print("\n⚠️ Some checks failed. Review issues before manual verification.")
            return 1
            
    except Exception as e:
        print(f"❌ Error checking Excel file: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main()) 